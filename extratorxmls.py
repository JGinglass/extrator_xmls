import os, glob
import pandas as pd
from lxml import etree
from decimal import Decimal, InvalidOperation
from openpyxl.utils import get_column_letter

NS_NFE = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

MONEY_COLS_FLAT = [
  
    "vUnCom",
    "vProd_item", "vProd_lote",
    "vICMS_item", "vICMS_lote",
    "vIPI_item", "vIPI_lote",
    "vPIS_item", "vPIS_lote",
    "vCOFINS_item", "vCOFINS_lote",

   
    "vBCST_item", "vBCST_lote",
    "vICMSST_item", "vICMSST_lote",

   
    "vFCP_item", "vFCP_lote",
    "vBCFCP_item", "vBCFCP_lote",

   
    "vFCPST_item", "vFCPST_lote",
    "vBCFCPST_item", "vBCFCPST_lote",

   
    "vProd_total", "vNF",
    "vICMS_total", "vIPI_total", "vPIS_total", "vCOFINS_total",
    "vBCST_total", "vST_total", "vFCP_total", "vFCPST_total",
]
QTY_COLS_FLAT = ["qCom_item", "qCom"]
DATE_COLS_FLAT = ["dFab", "dVal"]

PCT_COLS_FLAT = [
    # ICMS normal
    "pICMS_item", "pRedBC_item",
    # ICMS ST
    "pMVAST_item", "pICMSST_item", "pRedBCST_item",
    # FCP / FCP ST
    "pFCP_item", "pFCPST_item",
]


def first_text(node, xpath, ns=NS_NFE):
    if node is None:
        return None
    el = node.find(xpath, namespaces=ns)
    return el.text.strip() if el is not None and el.text else None


def xpath_first_text(node, expr, ns=NS_NFE):
    """Usa xpath (suporta local-name()) e retorna texto do primeiro match."""
    if node is None:
        return None
    res = node.xpath(expr, namespaces=ns)
    if not res:
        return None
    x = res[0]
    if isinstance(x, etree._Element):
        return x.text.strip() if x.text else None
    s = str(x).strip()
    return s if s else None


def to_decimal(x):
    if x is None:
        return None
    if isinstance(x, Decimal):
        return x
    if isinstance(x, (int, float)):
        return Decimal(str(x))
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def quantize_money(d: Decimal | None):
    if d is None:
        return None
    return d.quantize(Decimal("0.01"))


def quantize_qty(d: Decimal | None):
    if d is None:
        return None
    return d.quantize(Decimal("0.000"))


def detect_infNFe(root):
    return root.find(".//nfe:infNFe", namespaces=NS_NFE)


def safe_rateio(valor_item: Decimal | None, q_lote: Decimal | None, q_item: Decimal | None):
    if valor_item is None or q_lote is None or q_item is None:
        return None
    if q_item == 0:
        return None
    return quantize_money(valor_item * (q_lote / q_item))


def format_columns_as_currency(ws, df, col_names):
    col_index = {name: i + 1 for i, name in enumerate(df.columns)}
    moeda_fmt = '[$-416]"R$" #,##0.00'
    for name in col_names:
        if name not in col_index:
            continue
        idx = col_index[name]
        col_letter = get_column_letter(idx)
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value is None:
                continue
            cell.number_format = moeda_fmt


def format_columns(ws, df, col_names, number_format):
    col_index = {name: i + 1 for i, name in enumerate(df.columns)}
    for name in col_names:
        if name not in col_index:
            continue
        idx = col_index[name]
        col_letter = get_column_letter(idx)
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value is None:
                continue
            cell.number_format = number_format


def to_numeric_cols(df, cols):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def to_date_cols(df, cols):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce").dt.date
    return df


def duplicatas_txt_from_infNFe(infNFe):
    """
    Monta string com duplicatas:
    'nDup|dVenc|vDup; nDup|dVenc|vDup; ...'
    """
    if infNFe is None:
        return None
    dups = infNFe.findall(".//nfe:cobr/nfe:dup", namespaces=NS_NFE)
    if not dups:
        return None
    parts = []
    for d in dups:
        nDup = first_text(d, "nfe:nDup")
        dVenc = first_text(d, "nfe:dVenc")
        vDup = first_text(d, "nfe:vDup")
        parts.append(f"{nDup or ''}|{dVenc or ''}|{vDup or ''}")
    return "; ".join(parts)


def parse_file_flat(path):
    tree = etree.parse(path)
    root = tree.getroot()

    infNFe = detect_infNFe(root)
    if infNFe is None:
        return []

    id_attr = infNFe.get("Id")
    chave = id_attr[3:] if id_attr and id_attr.startswith("NFe") else \
        root.findtext(".//nfe:protNFe/nfe:infProt/nfe:chNFe", namespaces=NS_NFE)

    ide = infNFe.find("nfe:ide", namespaces=NS_NFE)
    emit = infNFe.find("nfe:emit", namespaces=NS_NFE)
    dest = infNFe.find("nfe:dest", namespaces=NS_NFE)
    total = infNFe.find("nfe:total/nfe:ICMSTot", namespaces=NS_NFE)

    duplicatas_txt = duplicatas_txt_from_infNFe(infNFe)

    cab = {
        "arquivo": os.path.basename(path),
        "chave": chave,
        "mod": first_text(ide, "nfe:mod"),
        "serie": first_text(ide, "nfe:serie"),
        "nNF": first_text(ide, "nfe:nNF"),
        "dhEmi": first_text(ide, "nfe:dhEmi") or first_text(ide, "nfe:dEmi"),
        "tpAmb": first_text(ide, "nfe:tpAmb"),
        "natOp": first_text(ide, "nfe:natOp"),

        "emit_CNPJ": first_text(emit, "nfe:CNPJ"),
        "emit_xNome": first_text(emit, "nfe:xNome"),
        "emit_UF": first_text(emit, "nfe:enderEmit/nfe:UF"),
        "dest_doc": first_text(dest, "nfe:CNPJ") or first_text(dest, "nfe:CPF"),
        "dest_xNome": first_text(dest, "nfe:xNome"),
        "dest_UF": first_text(dest, "nfe:enderDest/nfe:UF"),

        "duplicatas_txt": duplicatas_txt,

        "vProd_total": quantize_money(to_decimal(first_text(total, "nfe:vProd"))),
        "vNF": quantize_money(to_decimal(first_text(total, "nfe:vNF"))),
        "vICMS_total": quantize_money(to_decimal(first_text(total, "nfe:vICMS"))),
        "vIPI_total": quantize_money(to_decimal(first_text(total, "nfe:vIPI"))),
        "vPIS_total": quantize_money(to_decimal(first_text(total, "nfe:vPIS"))),
        "vCOFINS_total": quantize_money(to_decimal(first_text(total, "nfe:vCOFINS"))),

        "vBCST_total": quantize_money(to_decimal(first_text(total, "nfe:vBCST"))),
        "vST_total": quantize_money(to_decimal(first_text(total, "nfe:vST"))),
        "vFCP_total": quantize_money(to_decimal(first_text(total, "nfe:vFCP"))),
        "vFCPST_total": quantize_money(to_decimal(first_text(total, "nfe:vFCPST"))),

        "cStat": root.findtext(".//nfe:protNFe/nfe:infProt/nfe:cStat", namespaces=NS_NFE),
        "xMotivo": root.findtext(".//nfe:protNFe/nfe:infProt/nfe:xMotivo", namespaces=NS_NFE),
    }

    linhas = []

    for det in infNFe.findall("nfe:det", namespaces=NS_NFE):
        prod = det.find("nfe:prod", namespaces=NS_NFE)
        imposto = det.find("nfe:imposto", namespaces=NS_NFE)

        nItem = int(det.get("nItem")) if det.get("nItem") else None

        cProd = first_text(prod, "nfe:cProd")
        xProd = first_text(prod, "nfe:xProd")

        cProdANVISA = first_text(prod, "nfe:med/nfe:cProdANVISA")

        qCom_item = quantize_qty(to_decimal(first_text(prod, "nfe:qCom")))
        vUnCom = quantize_money(to_decimal(first_text(prod, "nfe:vUnCom")))
        vProd_item = quantize_money(to_decimal(first_text(prod, "nfe:vProd")))

        vICMS_item = quantize_money(to_decimal(first_text(imposto, ".//nfe:ICMS//nfe:vICMS")))
        vIPI_item = quantize_money(to_decimal(first_text(imposto, ".//nfe:IPI//nfe:vIPI")))
        vPIS_item = quantize_money(to_decimal(first_text(imposto, ".//nfe:PIS//nfe:vPIS")))
        vCOFINS_item = quantize_money(to_decimal(first_text(imposto, ".//nfe:COFINS//nfe:vCOFINS")))

        pICMS_item = to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="pICMS"]'))
        pRedBC_item = to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="pRedBC"]'))

       
        vBCST_item = quantize_money(to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="vBCST"]')))
        pICMSST_item = to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="pICMSST"]'))
        vICMSST_item = quantize_money(to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="vICMSST"]')))
        pMVAST_item = to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="pMVAST"]'))
        pRedBCST_item = to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="pRedBCST"]'))

        # Preço real unitário considerando ICMS-ST: (vProd_item + vICMSST_item) / qCom_item.
        # ICMS-ST ausente é tratado como zero para que o campo continue calculável.
        if vProd_item is not None and qCom_item is not None and qCom_item != 0:
            _icmsst = vICMSST_item if vICMSST_item is not None else Decimal("0")
            vUnReal_item = ((vProd_item + _icmsst) / qCom_item).quantize(Decimal("0.0001"))
        else:
            vUnReal_item = None

        
        vBCFCP_item = quantize_money(to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="vBCFCP"]')))
        pFCP_item = to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="pFCP"]'))
        vFCP_item = quantize_money(to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="vFCP"]')))

     
        vBCFCPST_item = quantize_money(to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="vBCFCPST"]')))
        pFCPST_item = to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="pFCPST"]'))
        vFCPST_item = quantize_money(to_decimal(xpath_first_text(imposto, './/nfe:ICMS//*[local-name()="vFCPST"]')))

        rastros = prod.findall("nfe:rastro", namespaces=NS_NFE) if prod is not None else []

        base_item = {
            **cab,
            "nItem": nItem,
            "cProd": cProd,
            "xProd": xProd,
            "cProdANVISA": cProdANVISA,

            "NCM": first_text(prod, "nfe:NCM"),
            "CFOP": first_text(prod, "nfe:CFOP"),
            "uCom": first_text(prod, "nfe:uCom"),

            "qCom_item": qCom_item,

            "vUnCom": vUnCom,
            "vUnReal_item": vUnReal_item,
            "vProd_item": vProd_item,

            "vICMS_item": vICMS_item,
            "vIPI_item": vIPI_item,
            "vPIS_item": vPIS_item,
            "vCOFINS_item": vCOFINS_item,

            "pICMS_item": pICMS_item,
            "pRedBC_item": pRedBC_item,

            "vBCST_item": vBCST_item,
            "pICMSST_item": pICMSST_item,
            "vICMSST_item": vICMSST_item,
            "pMVAST_item": pMVAST_item,
            "pRedBCST_item": pRedBCST_item,

            "vBCFCP_item": vBCFCP_item,
            "pFCP_item": pFCP_item,
            "vFCP_item": vFCP_item,

            "vBCFCPST_item": vBCFCPST_item,
            "pFCPST_item": pFCPST_item,
            "vFCPST_item": vFCPST_item,
        }

       
        if not rastros:
            linhas.append({
                **base_item,

                "qCom": qCom_item,

                "vProd_lote": None,
                "vICMS_lote": None,
                "vIPI_lote": None,
                "vPIS_lote": None,
                "vCOFINS_lote": None,

                "vBCST_lote": None,
                "vICMSST_lote": None,

                "vBCFCP_lote": None,
                "vFCP_lote": None,

                "vBCFCPST_lote": None,
                "vFCPST_lote": None,

                "seqRastro": None,
                "nLote": None,
                "dFab": None,
                "dVal": None,
                "cAgreg": None,
            })
            continue

 
        seq = 0
        for r in rastros:
            seq += 1
            qLote = quantize_qty(to_decimal(first_text(r, "nfe:qLote")))

            linhas.append({
                **base_item,

                "qCom": qLote,

                "vProd_lote": safe_rateio(vProd_item, qLote, qCom_item),
                "vICMS_lote": safe_rateio(vICMS_item, qLote, qCom_item),
                "vIPI_lote": safe_rateio(vIPI_item, qLote, qCom_item),
                "vPIS_lote": safe_rateio(vPIS_item, qLote, qCom_item),
                "vCOFINS_lote": safe_rateio(vCOFINS_item, qLote, qCom_item),

                "vBCST_lote": safe_rateio(vBCST_item, qLote, qCom_item),
                "vICMSST_lote": safe_rateio(vICMSST_item, qLote, qCom_item),

                "vBCFCP_lote": safe_rateio(vBCFCP_item, qLote, qCom_item),
                "vFCP_lote": safe_rateio(vFCP_item, qLote, qCom_item),

                "vBCFCPST_lote": safe_rateio(vBCFCPST_item, qLote, qCom_item),
                "vFCPST_lote": safe_rateio(vFCPST_item, qLote, qCom_item),

                "seqRastro": seq,
                "nLote": first_text(r, "nfe:nLote"),
                "dFab": first_text(r, "nfe:dFab"),
                "dVal": first_text(r, "nfe:dVal"),
                "cAgreg": first_text(r, "nfe:cAgreg"),
            })

    return linhas


def main():
    pasta = "xmls"
    flat_list = []

    for path in glob.glob(os.path.join(pasta, "*.xml")):
        linhas = parse_file_flat(path)
        if not linhas:
            print(f"[PULADO] Não achei infNFe em: {path}")
            continue
        print(f"[OK] {os.path.basename(path)} linhas={len(linhas)}")
        flat_list.extend(linhas)

    df = pd.DataFrame(flat_list)

  
    df = to_numeric_cols(df, MONEY_COLS_FLAT + QTY_COLS_FLAT + ["nItem", "seqRastro", "vUnReal_item"] + PCT_COLS_FLAT)

    df = to_date_cols(df, DATE_COLS_FLAT)

    out_xlsx = "nfe_export.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="tabela", index=False)
        ws = writer.book["tabela"]

        format_columns_as_currency(ws, df, MONEY_COLS_FLAT)
        format_columns(ws, df, ["vUnReal_item"], '[$-416]"R$" #,##0.0000')
        format_columns(ws, df, QTY_COLS_FLAT, "[$-416]#,##0.000")
        format_columns(ws, df, ["nItem", "seqRastro"], "0")
        format_columns(ws, df, DATE_COLS_FLAT, "dd/mm/yyyy")

        format_columns(ws, df, PCT_COLS_FLAT, "[$-416]0.00")

    print(f"Gerado: {out_xlsx} (aba: tabela).")


if __name__ == "__main__":
    main()